# ============================================================================
#  DASHBOARD DE CONTROL DE COMPRAS — PLANTA QUMA / LYON AG
# ----------------------------------------------------------------------------
#  Herramienta de un solo uso. Al ejecutar la celda:
#    1. Pide subir uno o dos archivos:
#         • COMPRAS del SAE                       — obligatorio
#         • Proveedores_a_Clasificar.xlsx ya editado — OPCIONAL
#    2. Detecta los meses presentes en el SAE y te pide elegir cuáles incluir
#       en el reporte (uno, varios, rangos o todos).
#    3. Genera y descarga automáticamente:
#         • Dashboard_Compras_QUMA.html   — Reporte interactivo (Plotly)
#         • Proveedores_a_Clasificar.xlsx — Listado de TODOS los proveedores
#                                            (no se filtra por mes — preserva
#                                            tu trabajo de clasificación).
#
#  El script auto-detecta cuál archivo es cuál inspeccionando su estructura,
#  así que puedes subir uno o ambos en el mismo diálogo.
#
#  Prioridad de categorización:
#    (1) Categoría validada por usuario en el Excel  →  GANA
#    (2) Ancla automática (7 reglas con evidencia)   →  fallback
#    (3) "Pendiente clasificar"                       →  para revisar manual
# ============================================================================

# ---------- 0. CONFIGURACIÓN (variables editables) --------------------------
HOJA_SAE             = "2026"
ESTATUS_VALIDOS      = ["Emitida", "Dev.Parc.", "Dev. Parcial"]
TOP_N_PROVEEDORES    = 10
TOP_N_FACTURAS       = 10
TOP_N_SIN_CLASIFICAR = 15

NOMBRE_HTML  = "Dashboard_Compras_QUMA.html"
NOMBRE_XLSX  = "Proveedores_a_Clasificar.xlsx"

# Catálogo cerrado de 11 categorías (acordado con el usuario).
CATALOGO_CATEGORIAS = [
    "Sustratos (Papel)",
    "Pre-prensa y Químicos",
    "Encuadernación",
    "Insumos de Producción",
    "Mantenimiento y Refacciones",
    "Maquila",
    "Logística / Fletes",
    "Almacenaje y Renta",
    "Limpieza y Sanitarios",
    "Servicios Profesionales",
    "Otros / Sin clasificar",
]

# Etiqueta interna para los proveedores que aún no han sido clasificados.
# NO es parte del catálogo del dropdown — es solo un estado interno del reporte.
ETIQ_PENDIENTE = "Pendiente clasificar"

# Anclas: las únicas reglas con evidencia documental fuerte (hoja TOP 10 del SAE).
# El resto se etiqueta como "Pendiente clasificar" y se valida en iteración 2.
def aplicar_ancla(proveedor):
    if pd.isna(proveedor):
        return None
    p = str(proveedor).upper()
    if "DELMAN INTERNACIONAL" in p:                          return "Sustratos (Papel)"
    if "SANCHEZ S.A"          in p:                          return "Pre-prensa y Químicos"
    if "LIBER ARTS"           in p:                          return "Encuadernación"
    if "JAQUELINA REYES"      in p or "GUTMAN BROS" in p:    return "Mantenimiento y Refacciones"
    if "VIGMAN GRAPHICS"      in p:                          return "Almacenaje y Renta"
    if "INFOVITA"             in p:                          return "Maquila"
    return None

# ---------- 1. IMPORTS E INSTALACIONES --------------------------------------
import sys, subprocess, io, os, re, warnings, datetime as dt
warnings.filterwarnings("ignore")

def _ensure(spec, import_name=None):
    """Instala silenciosamente un paquete si no está disponible."""
    try:
        __import__(import_name or spec.split("==")[0])
    except ImportError:
        subprocess.run([sys.executable, "-m", "pip", "install", "-q", spec], check=True)

# Sólo Plotly (interactivo) + openpyxl (Excel) son necesarios. matplotlib y
# reportlab fueron eliminados al quitar la generación del PDF: el script ahora
# es más liviano y no depende de binarios externos como Chrome o kaleido.
for spec, imp in [("plotly", "plotly"),
                  ("openpyxl", "openpyxl")]:
    _ensure(spec, imp)

import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---------- 2. SUBIDA DE ARCHIVOS (SAE obligatorio + clasificado opcional) --
try:
    from google.colab import files
except ImportError:
    raise EnvironmentError(
        "Este script está pensado para Google Colab. En otro entorno, define "
        "manualmente las rutas y omite la subida interactiva."
    )

print("📂 Sube uno o dos archivos:")
print("   • COMPRAS del SAE   (obligatorio)")
print("   • Proveedores_a_Clasificar.xlsx  (opcional — solo si ya lo trabajaste)")
print("   Tip: en el diálogo puedes seleccionar varios archivos con Ctrl/Cmd-clic.\n")
uploaded = files.upload()
if not uploaded:
    raise RuntimeError("No se subió ningún archivo. Vuelve a ejecutar la celda.")

# Auto-detección por estructura. El SAE tiene la hoja transaccional con la
# columna 'Tipo de cambio' en la fila 2; el clasificado tiene la hoja
# 'Clasificación' con la columna 'Proveedor_Exacto_SAE'.
def _identificar(nombre, contenido):
    if not nombre.lower().endswith((".xlsx", ".xlsm")):
        return None
    try:
        xls = pd.ExcelFile(io.BytesIO(contenido))
        if "Clasificación" in xls.sheet_names:
            # Probamos lectura con header en fila 1 (versión nueva con advertencia)
            # y en fila 0 (versión vieja). Si alguna tiene la columna llave, es clasificado.
            for hdr in (1, 0):
                try:
                    df_cls = pd.read_excel(io.BytesIO(contenido), sheet_name="Clasificación", header=hdr)
                    cols = [re.sub(r"[🔒✏️⚠️\s]+", "", str(c)) for c in df_cls.columns]
                    if any("Proveedor_Exacto_SAE" in c for c in cols):
                        return "clasificado"
                except Exception:
                    continue
        # Estrategia para SAE: detectar TODAS las hojas válidas y quedarnos
        # con (1) la hoja configurada en HOJA_SAE si existe, o (2) la más grande.
        candidatos = []
        for sheet in xls.sheet_names:
            try:
                df_test = pd.read_excel(io.BytesIO(contenido), sheet_name=sheet, header=1)
                cols = [str(c).strip() for c in df_test.columns]
                if "Proveedor" in cols and "Tipo de cambio" in cols and "Estatus" in cols:
                    candidatos.append((sheet, len(df_test)))
            except Exception:
                continue
        if candidatos:
            # Prioridad 1: la hoja configurada por el usuario, si está entre los candidatos.
            for sheet, _ in candidatos:
                if sheet == HOJA_SAE:
                    return ("sae", sheet)
            # Prioridad 2: la hoja con más filas (la "principal").
            mejor = max(candidatos, key=lambda x: x[1])
            return ("sae", mejor[0])
        return None
    except Exception:
        return None

ARCHIVO_SAE = bytes_sae = HOJA_SAE_DETECTADA = None
ARCHIVO_CLASIF = bytes_clasif = None

for nombre, contenido in uploaded.items():
    tipo = _identificar(nombre, contenido)
    if tipo == "clasificado":
        ARCHIVO_CLASIF, bytes_clasif = nombre, contenido
    elif isinstance(tipo, tuple) and tipo[0] == "sae":
        ARCHIVO_SAE, bytes_sae, HOJA_SAE_DETECTADA = nombre, contenido, tipo[1]

if ARCHIVO_SAE is None:
    raise RuntimeError(
        "No se identificó un archivo SAE válido entre los subidos.\n"
        f"Archivos recibidos: {list(uploaded)}\n"
        "El SAE debe tener una hoja con columnas 'Proveedor', 'Estatus' y "
        "'Tipo de cambio' en su segunda fila."
    )

# Si el usuario configuró una hoja específica y existe, respétala; si no, usa la detectada.
if HOJA_SAE_DETECTADA != HOJA_SAE:
    print(f"ℹ️  La hoja detectada del SAE es '{HOJA_SAE_DETECTADA}' "
          f"(el script estaba configurado para '{HOJA_SAE}'). Usaré la detectada.")
    HOJA_SAE = HOJA_SAE_DETECTADA

print(f"✅ SAE:           {ARCHIVO_SAE}  ({len(bytes_sae):,} bytes, hoja '{HOJA_SAE}')")
if bytes_clasif is not None:
    print(f"✅ Clasificación: {ARCHIVO_CLASIF}  ({len(bytes_clasif):,} bytes)")
else:
    print(f"ℹ️  Sin archivo de clasificación previa — el reporte usará solo las "
          f"anclas automáticas.")
print()

# ---------- 3. CARGA Y LIMPIEZA (ETL) ---------------------------------------
# El encabezado real del SAE está en la SEGUNDA fila (header=1).
try:
    df_raw = pd.read_excel(io.BytesIO(bytes_sae), sheet_name=HOJA_SAE, header=1)
except ValueError as e:
    hojas = pd.ExcelFile(io.BytesIO(bytes_sae)).sheet_names
    raise ValueError(
        f"La hoja '{HOJA_SAE}' no existe. Hojas disponibles: {hojas}.\n"
        f"Edita HOJA_SAE al inicio del script. Detalle: {e}"
    )

# El SAE trae espacios al final de las cabeceras ('Estatus ', 'Tipo de cambio '…).
df_raw.columns = df_raw.columns.astype(str).str.strip()
df_raw = df_raw.loc[:, ~df_raw.columns.str.startswith("Unnamed")]

COLS_REQ = ["Proveedor", "Estatus", "Referencia factura",
            "Fecha de documento", "Tipo de cambio", "Total del documento"]
faltantes = [c for c in COLS_REQ if c not in df_raw.columns]
if faltantes:
    raise KeyError(f"Faltan columnas en el SAE: {faltantes}. "
                   f"Detectadas: {list(df_raw.columns)}")

# Filtro de estatus: conservar Emitida + devolución parcial; descartar canceladas.
df = df_raw[df_raw["Estatus"].astype(str).str.strip().isin(ESTATUS_VALIDOS)].copy()
descartadas = len(df_raw) - len(df)

# Conversión defensiva de fechas y montos.
for col in ["Fecha de documento", "Fecha de recepción"]:
    if col in df.columns:
        df[col] = pd.to_datetime(df[col], errors="coerce")
df["Tipo de cambio"]      = pd.to_numeric(df["Tipo de cambio"], errors="coerce").fillna(1.0)
df["Total del documento"] = pd.to_numeric(df["Total del documento"], errors="coerce").fillna(0.0)
df["Gasto_Total_MXN"]     = df["Total del documento"] * df["Tipo de cambio"]

df = df.dropna(subset=["Fecha de documento"]).copy()
print(f"📥 SAE limpio: {len(df):,} facturas válidas "
      f"({descartadas} canceladas/nulas descartadas)")

# ---------- 3b. SELECCIÓN INTERACTIVA DE MESES ------------------------------
# Detectamos los meses presentes en el SAE y le preguntamos al usuario cuáles
# incluir en el reporte. El filtro afecta TODO el reporte (gráficas, KPIs,
# tabla de facturas), pero NO al Excel de clasificación: éste siempre incluye
# todos los proveedores del SAE original para preservar tu trabajo manual.
df["_Mes"] = df["Fecha de documento"].dt.to_period("M")
meses_disponibles = sorted(df["_Mes"].unique())
ESPANOL_MES = {1:"Ene", 2:"Feb", 3:"Mar", 4:"Abr", 5:"May", 6:"Jun",
               7:"Jul", 8:"Ago", 9:"Sep", 10:"Oct", 11:"Nov", 12:"Dic"}
def _label_mes(p):
    return f"{ESPANOL_MES[p.month]} {p.year}"

print("\n📅 Meses detectados en el SAE:")
for i, mes in enumerate(meses_disponibles, 1):
    sub = df[df["_Mes"] == mes]
    print(f"     {i}. {_label_mes(mes):12}  ·  {len(sub):>5} facturas  ·  "
          f"${sub['Gasto_Total_MXN'].sum():>14,.2f} MXN")

if len(meses_disponibles) == 1:
    # Un solo mes: no hay decisión real, lo informamos y seguimos.
    meses_seleccionados = list(meses_disponibles)
    print(f"\n   → Sólo hay un mes en el archivo, no hay selección por hacer.")
else:
    print(f"\n¿Qué meses quieres incluir en el reporte?")
    print(f"     • Números separados por coma (ej: 1,3,5)")
    print(f"     • Rangos con guion         (ej: 1-3)")
    print(f"     • Combinación              (ej: 1,3-5)")
    print(f"     • 'todos' o ENTER vacío    → incluir todos")

    def _parsear_seleccion(texto, n):
        """'1,3-5' → [1,3,4,5]. Tolerante a espacios y texto extra."""
        indices = set()
        for parte in texto.replace(" ", "").split(","):
            if not parte:
                continue
            if "-" in parte:
                a, b = parte.split("-", 1)
                a, b = int(a), int(b)
                indices.update(range(min(a, b), max(a, b) + 1))
            else:
                indices.add(int(parte))
        return sorted([i for i in indices if 1 <= i <= n])

    while True:
        try:
            sel = input("\n   Tu selección: ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            sel = "todos"
        if not sel or sel in ("todos", "all", "todo", "*"):
            meses_seleccionados = list(meses_disponibles)
            break
        try:
            idx = _parsear_seleccion(sel, len(meses_disponibles))
            if not idx:
                print(f"   ⚠️  Ninguna opción válida en '{sel}'. Intenta de nuevo.")
                continue
            meses_seleccionados = [meses_disponibles[i-1] for i in idx]
            break
        except Exception as e:
            print(f"   ⚠️  No entendí '{sel}' ({e}). Intenta con '1,3' o '1-3' o 'todos'.")

# Snapshot del SAE COMPLETO (todos los meses) — se usa solo para regenerar el
# Excel de clasificación, así nunca se pierde un proveedor por filtrar meses.
df_full = df.copy()

# Aplicar el filtro al df que alimenta el reporte.
df = df[df["_Mes"].isin(meses_seleccionados)].copy()

etiquetas_sel = ", ".join(_label_mes(m) for m in meses_seleccionados)
print(f"\n✅ Reporte incluirá: {etiquetas_sel}")
print(f"   Facturas filtradas: {len(df):,} de {len(df_full):,}  "
      f"({len(df)/len(df_full)*100:.1f}%)")
if len(df) == 0:
    raise RuntimeError("El filtro de meses no dejó ninguna factura. Aborto.")

# ---------- 4. CATEGORIZACIÓN ----------------------------------------------
# Prioridad:  (1) Categoría Validada por el usuario en el Excel clasificado
#             (2) Ancla automática (7 reglas con evidencia documental)
#             (3) "Pendiente clasificar"
#
# El humano siempre gana. Si el usuario asignó algo en el Excel previo, esa
# decisión sobrescribe la ancla automática.

# Mapas que vienen del Excel clasificado (vacíos si no se subió).
mapa_usuario_cat   = {}   # proveedor_exacto → categoría validada
mapa_usuario_notas = {}   # proveedor_exacto → notas
categorias_invalidas = []

if bytes_clasif is not None:
    print(f"📋 Leyendo clasificaciones del usuario desde '{ARCHIVO_CLASIF}'…")
    try:
        # Versiones anteriores del script ponían los headers en fila 1; la
        # versión actual los pone en fila 2 (con una fila de advertencia arriba).
        # Probamos ambas y nos quedamos con la que tenga las columnas esperadas.
        df_clasif = None
        for hdr in (1, 0):
            tmp = pd.read_excel(io.BytesIO(bytes_clasif), sheet_name="Clasificación", header=hdr)
            tmp.columns = [str(c).strip() for c in tmp.columns]
            cols_limpias = [re.sub(r"[🔒✏️⚠️\s]+", "", c) for c in tmp.columns]
            if any("Proveedor_Exacto_SAE" in c for c in cols_limpias):
                # Renombrar las columnas a su forma canónica (quitando iconos)
                tmp.columns = cols_limpias
                # Re-mapeo a nombres oficiales
                rename = {}
                for c in tmp.columns:
                    if "Proveedor_Exacto_SAE" in c: rename[c] = "Proveedor_Exacto_SAE"
                    elif "CategoríaValidada" in c or "Categoría Validada" in c: rename[c] = "Categoría Validada"
                    elif c == "Notas": rename[c] = "Notas"
                tmp = tmp.rename(columns=rename)
                df_clasif = tmp
                break

        if df_clasif is None:
            raise ValueError("No se encontró la columna 'Proveedor_Exacto_SAE' en la hoja.")

        provs_sae = set(df["Proveedor"].dropna().astype(str).str.strip())

        for _, row in df_clasif.iterrows():
            prov  = row.get("Proveedor_Exacto_SAE")
            cat_v = row.get("Categoría Validada")
            notas = row.get("Notas")
            if pd.isna(prov) or not str(prov).strip():
                continue
            prov = str(prov).strip()
            # Categoría: válida solo si está en el catálogo.
            if pd.notna(cat_v) and str(cat_v).strip():
                cat_v = str(cat_v).strip()
                if cat_v in CATALOGO_CATEGORIAS:
                    mapa_usuario_cat[prov] = cat_v
                else:
                    categorias_invalidas.append((prov, cat_v))
            if pd.notna(notas) and str(notas).strip():
                mapa_usuario_notas[prov] = str(notas).strip()

        # Conteo separado: validados que SÍ están en el SAE actual
        provs_no_en_sae = [p for p in mapa_usuario_cat if p not in provs_sae]
        provs_aplicados = len(mapa_usuario_cat) - len(provs_no_en_sae)

        print(f"   ↳ {provs_aplicados} proveedores validados que aplican al SAE actual")
        if mapa_usuario_notas:
            n_notas_aplicadas = sum(1 for p in mapa_usuario_notas if p in provs_sae)
            print(f"   ↳ {n_notas_aplicadas} con notas")
        if provs_no_en_sae:
            print(f"   ⚠️  {len(provs_no_en_sae)} proveedor(es) del Excel ya no aparecen en el SAE:")
            for p in provs_no_en_sae[:5]:
                print(f"      · {p[:60]}")
            if len(provs_no_en_sae) > 5:
                print(f"      … y {len(provs_no_en_sae)-5} más (probablemente dejaron de facturar)")
        if categorias_invalidas:
            print(f"   ⚠️  {len(categorias_invalidas)} celdas con categoría fuera del "
                  f"catálogo — serán ignoradas:")
            for prov, cat in categorias_invalidas[:5]:
                print(f"      · {prov[:50]:50}  →  '{cat}'")
            if len(categorias_invalidas) > 5:
                print(f"      … y {len(categorias_invalidas)-5} más")
    except Exception as e:
        print(f"   ⚠️  No se pudo leer el archivo clasificado: {e}")
        print(f"   ↳ Continuando solo con anclas automáticas.")

def categorizar(proveedor):
    """Usuario > Ancla > Pendiente"""
    if pd.isna(proveedor):
        return ETIQ_PENDIENTE
    p = str(proveedor).strip()
    if p in mapa_usuario_cat:
        return mapa_usuario_cat[p]
    ancla = aplicar_ancla(p)
    return ancla if ancla else ETIQ_PENDIENTE

def origen_categoria(proveedor):
    """Devuelve la fuente: 'usuario', 'ancla' o 'pendiente'."""
    if pd.isna(proveedor):
        return "pendiente"
    p = str(proveedor).strip()
    if p in mapa_usuario_cat:
        return "usuario"
    if aplicar_ancla(p):
        return "ancla"
    return "pendiente"

df["Categoria"] = df["Proveedor"].apply(categorizar)
df["Origen"]    = df["Proveedor"].apply(origen_categoria)

# Categorización también sobre el SAE completo (sin filtro de meses) para el Excel.
df_full["Categoria"] = df_full["Proveedor"].apply(categorizar)
df_full["Origen"]    = df_full["Proveedor"].apply(origen_categoria)

# ---------- 5. KPIs ---------------------------------------------------------
gasto_total      = df["Gasto_Total_MXN"].sum()
total_facturas   = len(df)
ticket_promedio  = df["Gasto_Total_MXN"].mean()
ticket_mediano   = df["Gasto_Total_MXN"].median()
fecha_min        = df["Fecha de documento"].min()
fecha_max        = df["Fecha de documento"].max()
dias_cubiertos   = (fecha_max - fecha_min).days + 1
prov_unicos      = df["Proveedor"].nunique()

mask_anclado     = df["Categoria"] != ETIQ_PENDIENTE
gasto_anclado    = df.loc[ mask_anclado, "Gasto_Total_MXN"].sum()
gasto_pendiente  = df.loc[~mask_anclado, "Gasto_Total_MXN"].sum()
prov_anclados    = df.loc[ mask_anclado, "Proveedor"].nunique()
prov_pendientes  = df.loc[~mask_anclado, "Proveedor"].nunique()
pct_cobertura    = gasto_anclado / gasto_total * 100 if gasto_total else 0

# Desglose por origen (sólo relevante si se proveyó el archivo clasificado)
gasto_por_origen = df.groupby("Origen")["Gasto_Total_MXN"].sum().to_dict()
prov_por_origen  = df.groupby("Origen")["Proveedor"].nunique().to_dict()

print("\n" + "═"*74)
print("  RESUMEN EJECUTIVO — CONTROL DE COMPRAS QUMA")
print("═"*74)
print(f"  Periodo                       : {fecha_min:%d-%b-%Y} → {fecha_max:%d-%b-%Y}  ({dias_cubiertos} días)")
print(f"  Gasto Total Acumulado         : ${gasto_total:>18,.2f} MXN")
print(f"  Facturas válidas              : {total_facturas:>22,}")
print(f"  Ticket promedio               : ${ticket_promedio:>18,.2f} MXN")
print(f"  Ticket mediano                : ${ticket_mediano:>18,.2f} MXN")
print(f"  Proveedores únicos            : {prov_unicos:>22,}")
print("─"*74)
print(f"  COBERTURA DE CATEGORIZACIÓN — {pct_cobertura:.1f}% del gasto")
if bytes_clasif is not None:
    g_user = gasto_por_origen.get("usuario", 0)
    g_anc  = gasto_por_origen.get("ancla", 0)
    p_user = prov_por_origen.get("usuario", 0)
    p_anc  = prov_por_origen.get("ancla", 0)
    print(f"    · Validado por usuario      : ${g_user:>16,.2f}  ({g_user/gasto_total*100:5.1f}%) — {p_user:3} provs")
    print(f"    · Ancla automática          : ${g_anc:>16,.2f}  ({g_anc/gasto_total*100:5.1f}%) — {p_anc:3} provs")
print(f"    · Pendiente clasificar        : ${gasto_pendiente:>16,.2f}  ({100-pct_cobertura:5.1f}%) — {prov_pendientes:3} provs")
print("═"*74)

# ---------- 6. PALETA DE COLORES --------------------------------------------
PALETA_BASE = ["#1F4E79", "#C00000", "#BF8F00", "#548235", "#7030A0",
               "#2E75B6", "#E97132", "#385723", "#806000", "#A02B93"]
PALETA = {cat: PALETA_BASE[i % len(PALETA_BASE)]
          for i, cat in enumerate(CATALOGO_CATEGORIAS[:-1])}  # excluye "Otros"
PALETA["Otros / Sin clasificar"] = "#9E9E9E"
PALETA[ETIQ_PENDIENTE] = "#9E9E9E"   # gris honesto

# ============================================================================
#  GRÁFICO 1 — DONUT
# ============================================================================
gasto_cat = (df.groupby("Categoria", as_index=False)["Gasto_Total_MXN"].sum()
               .sort_values("Gasto_Total_MXN", ascending=False))
if bytes_clasif is not None:
    subtit_donut = (f"Cobertura: <b>{pct_cobertura:.1f}%</b>  ·  "
                    f"{prov_pendientes} proveedores aún pendientes")
else:
    subtit_donut = (f"Cobertura anclada con evidencia documental: "
                    f"<b>{pct_cobertura:.1f}%</b>  ·  el segmento gris se valida "
                    f"manualmente en el Excel adjunto")

fig1 = px.pie(
    gasto_cat, names="Categoria", values="Gasto_Total_MXN",
    hole=0.55, color="Categoria", color_discrete_map=PALETA,
    title=("<b>Distribución del Gasto por Categoría</b>"
           f"<br><sup>{subtit_donut}</sup>"),
)
fig1.update_traces(textposition="outside", textinfo="percent+label", sort=False,
                   hovertemplate="<b>%{label}</b><br>Gasto: $%{value:,.0f} MXN<br>"
                                 "Participación: %{percent}<extra></extra>")
fig1.update_layout(
    template="plotly_white",
    annotations=[dict(text=f"<b>${gasto_total/1e6:,.1f}M</b><br>"
                            f"<span style='font-size:12px'>MXN total</span>",
                      x=0.5, y=0.5, font=dict(size=18), showarrow=False)],
    height=560, margin=dict(t=120, b=40, l=40, r=40),
    legend=dict(orientation="v", yanchor="middle", y=0.5),
)
fig1.show()

# ============================================================================
#  GRÁFICO 2 — CURVA DE GASTO SEMANAL
# ============================================================================
df_sem = (df.set_index("Fecha de documento")
            .resample("W-MON")["Gasto_Total_MXN"].sum().reset_index())
promedio_sem = df_sem["Gasto_Total_MXN"].mean()
pico_idx = df_sem["Gasto_Total_MXN"].idxmax()
pico_x   = df_sem.loc[pico_idx, "Fecha de documento"]
pico_y   = df_sem.loc[pico_idx, "Gasto_Total_MXN"]

fig2 = go.Figure()
fig2.add_trace(go.Scatter(
    x=df_sem["Fecha de documento"], y=df_sem["Gasto_Total_MXN"],
    mode="lines+markers", fill="tozeroy",
    line=dict(color="#1F4E79", width=2.5),
    fillcolor="rgba(31, 78, 121, 0.18)",
    marker=dict(size=7, color="#1F4E79"),
    hovertemplate="Semana del %{x|%d-%b-%Y}<br>Gasto: $%{y:,.0f} MXN<extra></extra>",
))
fig2.add_hline(y=promedio_sem, line_dash="dash", line_color="#C00000",
               annotation_text=f"Promedio semanal: ${promedio_sem:,.0f}",
               annotation_position="top right", annotation_font_color="#C00000")
fig2.add_annotation(x=pico_x, y=pico_y, text=f"<b>PICO</b><br>${pico_y/1e6:.2f}M",
                    showarrow=True, arrowhead=2, arrowcolor="#C00000",
                    font=dict(color="#C00000", size=11), yshift=12)
fig2.update_layout(
    title=("<b>Curva de Gasto Semanal — Detección de Picos</b>"
           "<br><sup>Agrupación dinámica L–D; se ajusta automáticamente a meses futuros</sup>"),
    xaxis_title="Semana", yaxis_title="Gasto (MXN)",
    template="plotly_white", height=480, showlegend=False,
    margin=dict(t=100, b=60, l=70, r=40),
)
fig2.update_yaxes(tickformat=",.0f", tickprefix="$")
fig2.show()

# ============================================================================
#  GRÁFICO 3 — PARETO TOP 10 PROVEEDORES
# ============================================================================
top_prov = (df.groupby("Proveedor", as_index=False)["Gasto_Total_MXN"].sum()
              .sort_values("Gasto_Total_MXN", ascending=False)
              .head(TOP_N_PROVEEDORES))
prov_to_cat = df.groupby("Proveedor")["Categoria"].first().to_dict()
top_prov["Categoria"] = top_prov["Proveedor"].map(prov_to_cat)
top_prov["Proveedor_Display"] = top_prov["Proveedor"].apply(
    lambda x: x if len(x) <= 38 else x[:35] + "…")
top_prov = top_prov.sort_values("Gasto_Total_MXN", ascending=False).reset_index(drop=True)
top_prov["Pct_Acumulado"] = top_prov["Gasto_Total_MXN"].cumsum() / gasto_total * 100
pct_top10 = top_prov["Pct_Acumulado"].iloc[-1]

fig3 = px.bar(
    top_prov.sort_values("Gasto_Total_MXN", ascending=True),
    x="Gasto_Total_MXN", y="Proveedor_Display", orientation="h",
    color="Categoria", color_discrete_map=PALETA, text="Gasto_Total_MXN",
    title=(f"<b>Pareto — Top {TOP_N_PROVEEDORES} Proveedores por Gasto (MXN)</b>"
           f"<br><sup>El top {TOP_N_PROVEEDORES} concentra el <b>{pct_top10:.1f}%</b> "
           f"del gasto total del periodo</sup>"),
)
fig3.update_traces(texttemplate="$%{text:,.0f}", textposition="outside",
                   hovertemplate="<b>%{y}</b><br>Gasto: $%{x:,.0f} MXN<extra></extra>")
fig3.update_layout(template="plotly_white", xaxis_title="Gasto (MXN)", yaxis_title="",
                   height=560, legend_title="Categoría",
                   margin=dict(t=110, b=60, l=20, r=120))
fig3.update_xaxes(tickformat=",.0f", tickprefix="$")
fig3.show()

# ============================================================================
#  GRÁFICO 4 — TOP PROVEEDORES PENDIENTES DE CLASIFICAR (lista de acción)
# ============================================================================
pend = (df[~mask_anclado].groupby("Proveedor", as_index=False)["Gasto_Total_MXN"].sum()
          .sort_values("Gasto_Total_MXN", ascending=False)
          .head(TOP_N_SIN_CLASIFICAR))
pend["Proveedor_Display"] = pend["Proveedor"].apply(
    lambda x: x if len(x) <= 40 else x[:37] + "…")

fig4 = None
if len(pend) > 0:
    fig4 = px.bar(
        pend.sort_values("Gasto_Total_MXN", ascending=True),
        x="Gasto_Total_MXN", y="Proveedor_Display", orientation="h",
        text="Gasto_Total_MXN",
        title=(f"<b>Top {TOP_N_SIN_CLASIFICAR} Proveedores Pendientes de Clasificar</b>"
               f"<br><sup>Clasifícalos en el Excel adjunto para subir la cobertura "
               f"desde el {pct_cobertura:.1f}% actual</sup>"),
        color_discrete_sequence=["#9E9E9E"],
    )
    fig4.update_traces(texttemplate="$%{text:,.0f}", textposition="outside",
                       hovertemplate="<b>%{y}</b><br>Gasto: $%{x:,.0f} MXN<extra></extra>")
    fig4.update_layout(template="plotly_white", xaxis_title="Gasto (MXN)", yaxis_title="",
                       height=560, showlegend=False,
                       margin=dict(t=110, b=60, l=20, r=120))
    fig4.update_xaxes(tickformat=",.0f", tickprefix="$")
    fig4.show()

# ============================================================================
#  TABLA — TOP 10 FACTURAS INDIVIDUALES
# ============================================================================
top_fac = (df.nlargest(TOP_N_FACTURAS, "Gasto_Total_MXN")
             [["Proveedor", "Fecha de documento", "Referencia factura",
               "Gasto_Total_MXN", "Categoria"]]
             .copy().reset_index(drop=True))
top_fac_fmt = top_fac.copy()
top_fac_fmt["Fecha de documento"] = top_fac_fmt["Fecha de documento"].dt.strftime("%d-%b-%Y")
top_fac_fmt["Gasto_Total_MXN"]    = top_fac_fmt["Gasto_Total_MXN"].apply(lambda x: f"${x:,.2f}")
top_fac_fmt.index = top_fac_fmt.index + 1
top_fac_fmt.columns = ["Proveedor", "Fecha", "Referencia", "Gasto (MXN)", "Categoría"]

fig5 = go.Figure(data=[go.Table(
    columnwidth=[30, 220, 80, 110, 130, 180],
    header=dict(values=["<b>#</b>","<b>Proveedor</b>","<b>Fecha</b>",
                        "<b>Referencia</b>","<b>Gasto (MXN)</b>","<b>Categoría</b>"],
                fill_color="#1F4E79", font=dict(color="white", size=12),
                align="left", height=34),
    cells=dict(values=[list(top_fac_fmt.index), top_fac_fmt["Proveedor"],
                       top_fac_fmt["Fecha"], top_fac_fmt["Referencia"],
                       top_fac_fmt["Gasto (MXN)"], top_fac_fmt["Categoría"]],
               fill_color=[["#F4F7FA","#FFFFFF"]*TOP_N_FACTURAS],
               align="left", font=dict(size=11), height=28),
)])
fig5.update_layout(
    title=("<b>Top 10 Facturas Individuales por Monto</b>"
           "<br><sup>Identifica inyecciones fuertes de capital (CapEx) y compras masivas</sup>"),
    height=460, margin=dict(t=90, b=20, l=10, r=10),
)
fig5.show()

print("\n" + "═"*74)
print("  TOP 10 FACTURAS DE MAYOR VALOR")
print("═"*74)
print(top_fac_fmt.to_string())
print("═"*74)

# ============================================================================
#  ENTREGABLE 1 — HTML INTERACTIVO
# ============================================================================
print("\n🌐 Generando HTML interactivo…")

def fig_html(fig, include_js=False):
    return pio.to_html(fig, full_html=False, include_plotlyjs="inline" if include_js else False,
                       config={"displayModeBar": True, "displaylogo": False})

kpi_rows = "".join(f"""
    <tr><td>{k}</td><td>{v}</td></tr>""" for k, v in [
    ("Periodo analizado", f"{fecha_min:%d-%b-%Y} → {fecha_max:%d-%b-%Y} ({dias_cubiertos} días)"),
    ("Gasto Total Acumulado", f"${gasto_total:,.2f} MXN"),
    ("Facturas válidas", f"{total_facturas:,}"),
    ("Ticket promedio", f"${ticket_promedio:,.2f} MXN"),
    ("Ticket mediano", f"${ticket_mediano:,.2f} MXN"),
    ("Proveedores únicos", f"{prov_unicos}"),
    ("Cobertura anclada", f"{pct_cobertura:.1f}%  (${gasto_anclado:,.0f})"),
    ("Pendiente de clasificar", f"{100-pct_cobertura:.1f}%  ({prov_pendientes} proveedores)"),
])

html = f"""<!DOCTYPE html>
<html lang="es">
<head>
  <meta charset="utf-8">
  <title>Dashboard Compras QUMA — Lyon AG</title>
  <style>
    body {{ font-family: -apple-system, Segoe UI, Helvetica, Arial, sans-serif;
            margin: 0; padding: 0; background: #F4F7FA; color: #1f2933; }}
    .wrap {{ max-width: 1280px; margin: 0 auto; padding: 24px; }}
    h1 {{ color: #1F4E79; font-size: 26px; margin: 0 0 4px; }}
    .meta {{ color: #5b6b7a; font-size: 13px; margin-bottom: 24px; }}
    .card {{ background: white; border-radius: 8px; padding: 20px;
             margin-bottom: 20px; box-shadow: 0 1px 3px rgba(0,0,0,0.08); }}
    .card h2 {{ color: #1F4E79; margin: 0 0 12px; font-size: 16px;
                border-bottom: 2px solid #1F4E79; padding-bottom: 8px; }}
    table.kpi {{ width: 100%; border-collapse: collapse; font-size: 14px; }}
    table.kpi td {{ padding: 8px 12px; border-bottom: 1px solid #E1E7EC; }}
    table.kpi td:first-child {{ font-weight: 600; color: #1F4E79;
                                background: #F4F7FA; width: 35%; }}
    .note {{ background: #FFF8E1; border-left: 4px solid #BF8F00;
             padding: 12px 16px; font-size: 13px; color: #5d4500;
             border-radius: 4px; margin-top: 12px; }}
    footer {{ text-align: center; color: #8a99a8; font-size: 12px;
              padding: 20px; }}
  </style>
</head>
<body>
  <div class="wrap">
    <h1>Dashboard de Control de Compras — Planta QUMA</h1>
    <p class="meta">Lyon AG  ·  Meses incluidos: <b>{etiquetas_sel}</b>  ·  
       Archivo: {ARCHIVO_SAE}  ·  Generado {dt.datetime.now():%d-%b-%Y %H:%M}</p>

    <div class="card">
      <h2>Resumen Ejecutivo</h2>
      <table class="kpi">{kpi_rows}</table>
      <div class="note">
        <b>Metodología:</b> sólo 7 patrones de proveedor están anclados con evidencia
        documental (DELMAN → Sustratos, Sanchez S.A → Pre-prensa, etc.). El resto del
        gasto está marcado como <i>Pendiente clasificar</i> hasta que se valide en el
        Excel adjunto. El reporte no inventa categorías.
      </div>
    </div>

    <div class="card"><h2>Distribución del Gasto por Categoría</h2>{fig_html(fig1, include_js=True)}</div>
    <div class="card"><h2>Curva de Gasto Semanal — Detección de Picos</h2>{fig_html(fig2)}</div>
    <div class="card"><h2>Pareto — Top {TOP_N_PROVEEDORES} Proveedores</h2>{fig_html(fig3)}</div>
"""
if fig4 is not None:
    html += f"""    <div class="card"><h2>Top {TOP_N_SIN_CLASIFICAR} Proveedores Pendientes de Clasificar</h2>{fig_html(fig4)}</div>
"""
html += f"""    <div class="card"><h2>Top 10 Facturas Individuales</h2>{fig_html(fig5)}</div>

    <footer>Generado automáticamente · Lyon AG · QUMA</footer>
  </div>
</body>
</html>"""

with open(NOMBRE_HTML, "w", encoding="utf-8") as f:
    f.write(html)
print(f"✅ HTML generado: {NOMBRE_HTML}  ({os.path.getsize(NOMBRE_HTML):,} bytes)")

# ============================================================================
#  ENTREGABLE 3 — EXCEL DE CLASIFICACIÓN MANUAL
# ============================================================================
#  Estructura:
#    Hoja 1: "Clasificación"  — tabla editable
#       Col A (#)                 — read-only
#       Col B (Proveedor_Exacto_SAE) — LLAVE, read-only, NUNCA editar
#       Col C (Gasto Total MXN)   — read-only, formato moneda
#       Col D (# Facturas)        — read-only
#       Col E (Categoría Sugerida)— read-only (pre-llenada para anclas)
#       Col F (Categoría Validada)— EDITABLE con dropdown del catálogo
#       Col G (Notas)             — EDITABLE libre
#    Hoja 2: "_Catalogo" (oculta) — fuente del dropdown
#    Hoja 3: "Instrucciones" — guía rápida
#
#  La protección de hoja se activa con contraseña vacía: evita ediciones
#  accidentales en las columnas-llave, pero el usuario puede desproteger
#  manualmente sin contraseña si lo necesita (Revisar → Desproteger hoja).
print("📊 Generando Excel de clasificación…")

# Resumen por proveedor sobre el SAE COMPLETO (no el filtrado por mes), para
# que el Excel siempre contenga el universo total de proveedores y preserve
# todas las clasificaciones aunque el reporte muestre sólo algunos meses.
prov_summary = (df_full.groupby("Proveedor")
                  .agg(Gasto_Total_MXN=("Gasto_Total_MXN", "sum"),
                       Num_Facturas=("Gasto_Total_MXN", "count"))
                  .reset_index()
                  .sort_values("Gasto_Total_MXN", ascending=False))
prov_summary["Categoria_Sugerida"] = prov_summary["Proveedor"].apply(aplicar_ancla)

wb = Workbook()

# --- Hoja: Instrucciones (primera para que sea lo primero que ve el usuario) ---
ws_inst = wb.active
ws_inst.title = "Instrucciones"

instr_lines = [
    ("Cómo usar este archivo", "h1"),
    ("", "p"),
    (f"Este Excel contiene los {df_full['Proveedor'].nunique()} proveedores "
     "únicos detectados en tu archivo SAE de compras (incluyendo todos los "
     "meses, no sólo los que filtraste para el reporte). Tu trabajo es "
     "validar / asignar la categoría correcta en la columna 'Categoría "
     "Validada' usando el menú desplegable.", "p"),
    ("", "p"),
    ("Pasos:", "h2"),
    ("1. Ve a la hoja 'Clasificación'.", "p"),
    ("2. Para cada proveedor, selecciona una categoría en la columna F "
     "(✏️ Categoría Validada). Haz clic en la celda y aparecerá una flecha "
     "con el menú desplegable.", "p"),
    ("3. Los colores te orientan:", "p"),
    ("   • Azul claro  → ya fue clasificado por ti en una corrida previa", "p"),
    ("   • Verde claro → sugerencia automática (ancla con evidencia documental)", "p"),
    ("   • Ámbar       → pendiente, requiere tu decisión", "p"),
    ("4. La columna G (✏️ Notas) es libre para que escribas observaciones.", "p"),
    ("5. Guarda el archivo. La próxima vez que corras el script, súbelo junto con "
     "el SAE de compras y tu trabajo será preservado y aplicado al reporte.", "p"),
    ("", "p"),
    ("Convención visual (no es protección técnica):", "h2"),
    ("• Columnas con fondo gris claro y candado 🔒 en el header → NO EDITAR. "
     "Son referencia automática que el script regenera en cada corrida.", "p"),
    ("• Columna B (🔒 Proveedor_Exacto_SAE, fondo rojo claro) → es la LLAVE "
     "que el script usa para identificar a cada proveedor. Si la modificas, "
     "ese proveedor se perderá del mapeo y el script lo tratará como nuevo.", "p"),
    ("• Columnas F y G con borde verde y lápiz ✏️ → ahí trabajas tú.", "p"),
    ("", "p"),
    ("Qué pasa si me equivoco:", "h2"),
    ("• Si editas un nombre en la columna B por accidente, el script de la "
     "próxima corrida lo detectará y te avisará — tu clasificación se ignora "
     "para ese proveedor, pero el reporte sigue corriendo sin romperse.", "p"),
    ("• Si escribes una categoría a mano fuera del menú desplegable, el script "
     "te lo reportará y la ignorará. Usa siempre el dropdown.", "p"),
    ("• Si insertas o eliminas filas, el orden cambia pero el cruce sigue "
     "funcionando porque la llave es el nombre del proveedor, no la posición.", "p"),
    ("", "p"),
    ("Catálogo de categorías:", "h2"),
] + [(f"• {c}", "p") for c in CATALOGO_CATEGORIAS]

styles_inst = {
    "h1": Font(bold=True, size=16, color="1F4E79"),
    "h2": Font(bold=True, size=12, color="1F4E79"),
    "p":  Font(size=11, color="333333"),
}
for i, (txt, style) in enumerate(instr_lines, 1):
    c = ws_inst.cell(row=i, column=1, value=txt)
    c.font = styles_inst[style]
    c.alignment = Alignment(wrap_text=True, vertical="top")
ws_inst.column_dimensions["A"].width = 110
ws_inst.sheet_view.showGridLines = False

# --- Hoja: _Catalogo (oculta, fuente del dropdown) ---
ws_cat = wb.create_sheet("_Catalogo")
for i, cat in enumerate(CATALOGO_CATEGORIAS, 1):
    ws_cat.cell(row=i, column=1, value=cat)
ws_cat.sheet_state = "hidden"

# --- Hoja: Clasificación (la tabla principal) ---
ws = wb.create_sheet("Clasificación")

# FILA 1: barra de advertencia ocupando todo el ancho.
# Más efectiva que la protección de hoja, que en openpyxl causa más problemas
# de los que resuelve: el usuario ve la advertencia ANTES de tocar nada.
msg_warn = ("⚠️  EDITA SOLO LAS COLUMNAS F (Categoría Validada) y G (Notas)  ·  "
            "Las columnas A–E son referencia del script — si las cambias, "
            "el cruce con el SAE puede romperse")
ws.merge_cells("A1:G1")
warn_cell = ws.cell(row=1, column=1, value=msg_warn)
warn_cell.font = Font(bold=True, color="9C0006", size=11)
warn_cell.fill = PatternFill("solid", start_color="FFE5E5")  # rojo muy claro
warn_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
warn_cell.border = Border(
    top=Side(style="medium", color="C00000"),
    bottom=Side(style="medium", color="C00000"),
    left=Side(style="medium", color="C00000"),
    right=Side(style="medium", color="C00000"),
)
ws.row_dimensions[1].height = 32

# FILA 2: headers. Las columnas read-only (A–E) llevan un candado en el título;
# las editables (F y G) llevan un lápiz para indicar dónde se trabaja.
headers = [
    ("#", "ro"),
    ("🔒 Proveedor_Exacto_SAE", "ro_key"),
    ("Gasto Total MXN", "ro"),
    ("# Facturas", "ro"),
    ("Categoría Sugerida", "ro"),
    ("✏️ Categoría Validada", "edit"),
    ("✏️ Notas", "edit"),
]
fill_hdr_ro     = PatternFill("solid", start_color="4A6B8A")  # azul-gris atenuado
fill_hdr_ro_key = PatternFill("solid", start_color="8B0000")  # rojo oscuro = LLAVE
fill_hdr_edit   = PatternFill("solid", start_color="1F7A1F")  # verde = editable
for col, (h, tipo) in enumerate(headers, 1):
    cell = ws.cell(row=2, column=col, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = (fill_hdr_ro_key if tipo == "ro_key"
                 else fill_hdr_edit if tipo == "edit"
                 else fill_hdr_ro)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws.row_dimensions[2].height = 30

# Data validation enlazada a la hoja _Catalogo
dv = DataValidation(
    type="list",
    formula1=f"=_Catalogo!$A$1:$A${len(CATALOGO_CATEGORIAS)}",
    allow_blank=True,
    showDropDown=False,
)
dv.error      = "Categoría inválida. Selecciona una opción del menú desplegable."
dv.errorTitle = "Categoría inválida"
dv.prompt     = f"Selecciona una de las {len(CATALOGO_CATEGORIAS)} categorías del catálogo."
dv.promptTitle = "Categorizar proveedor"
ws.add_data_validation(dv)

# Paleta para celdas (sin protección — solo señales visuales).
fill_user      = PatternFill("solid", start_color="DCE7F5")  # azul claro = validado por usuario
fill_ancla     = PatternFill("solid", start_color="E6F2D8")  # verde claro = ancla automática
fill_pend      = PatternFill("solid", start_color="FFF8E1")  # ámbar = pendiente
fill_ro_data   = PatternFill("solid", start_color="F0F2F5")  # gris claro = read-only data
fill_key_data  = PatternFill("solid", start_color="FCE4E4")  # rojo muy claro = llave (B)
font_ro        = Font(size=10, color="555555")               # texto atenuado = no editar
font_ro_key    = Font(size=10, color="8B0000", bold=True)    # rojo oscuro = llave
font_edit      = Font(size=10, color="1F4E79", bold=True)
font_notes     = Font(size=10, color="333333")

for i, r in enumerate(prov_summary.itertuples(index=False), start=3):  # datos desde fila 3
    ws.cell(row=i, column=1, value=i-2)
    ws.cell(row=i, column=2, value=r.Proveedor)
    ws.cell(row=i, column=3, value=float(r.Gasto_Total_MXN))
    ws.cell(row=i, column=4, value=int(r.Num_Facturas))
    ws.cell(row=i, column=5, value=r.Categoria_Sugerida if r.Categoria_Sugerida else "")

    # Categoría Validada: prioridad → (1) usuario, (2) ancla, (3) vacío
    cat_usuario = mapa_usuario_cat.get(r.Proveedor)
    if cat_usuario:
        val_validada, origen = cat_usuario, "usuario"
    elif r.Categoria_Sugerida:
        val_validada, origen = r.Categoria_Sugerida, "ancla"
    else:
        val_validada, origen = "", "pendiente"
    ws.cell(row=i, column=6, value=val_validada)
    ws.cell(row=i, column=7, value=mapa_usuario_notas.get(r.Proveedor, ""))

    # Columnas A, C, D, E (read-only de info)
    for col in (1, 3, 4, 5):
        c = ws.cell(row=i, column=col)
        c.font = font_ro
        c.fill = fill_ro_data
        c.alignment = Alignment(vertical="center",
                                horizontal="right" if col in (3, 4) else "left")
    ws.cell(row=i, column=3).number_format = '"$"#,##0.00'

    # Columna B (LLAVE: rojo claro + texto rojo oscuro)
    c = ws.cell(row=i, column=2)
    c.font = font_ro_key
    c.fill = fill_key_data
    c.alignment = Alignment(vertical="center", horizontal="left")

    # Columna F (Categoría Validada): color según origen
    fill_F = (fill_user if origen == "usuario"
              else fill_ancla if origen == "ancla"
              else fill_pend)
    c = ws.cell(row=i, column=6)
    c.font = font_edit
    c.fill = fill_F
    c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    c.border = Border(left=Side(style="thin", color="1F7A1F"),
                      right=Side(style="thin", color="1F7A1F"))

    # Columna G (Notas)
    c = ws.cell(row=i, column=7)
    c.font = font_notes
    c.fill = PatternFill()
    c.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    c.border = Border(right=Side(style="thin", color="1F7A1F"))

    # Dropdown sólo en la columna F
    dv.add(ws.cell(row=i, column=6))

# Anchos de columna
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 48
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 12
ws.column_dimensions["E"].width = 28
ws.column_dimensions["F"].width = 30
ws.column_dimensions["G"].width = 40
# Freeze panes en C3: deja fijas la fila 1 (advertencia) + fila 2 (headers)
# + columnas A-B (#, proveedor) cuando el usuario hace scroll.
ws.freeze_panes = "C3"
ws.sheet_view.showGridLines = False
ws.sheet_view.zoomScale = 110   # un poco más grande para mejor lectura

# NOTA: no se aplica protección de hoja. Las celdas A-E quedan visualmente
# "deshabilitadas" (gris claro + texto atenuado) y la columna B (llave) en
# rojo, pero técnicamente todas son editables. Si el usuario rompe la llave
# por error, el script de la siguiente corrida detecta el desajuste y avisa
# (los proveedores no encontrados se ignoran sin fallar).

wb.save(NOMBRE_XLSX)
print(f"✅ Excel generado: {NOMBRE_XLSX}  ({os.path.getsize(NOMBRE_XLSX):,} bytes)")

# ============================================================================
#  DESCARGA AUTOMÁTICA DE LOS TRES ARCHIVOS
# ============================================================================
print("\n⬇️  Iniciando descargas…")
for f_path in (NOMBRE_HTML, NOMBRE_XLSX):
    try:
        files.download(f_path)
        print(f"   ↳ {f_path}")
    except Exception as e:
        print(f"   ⚠️  {f_path} disponible en el panel de archivos de Colab. ({e})")

print("\n" + "═"*74)
print("  REPORTE GENERADO")
print("═"*74)
print(f"  → Dashboard interactivo : {NOMBRE_HTML}")
print(f"  → Clasificación         : {NOMBRE_XLSX}")
if bytes_clasif is not None:
    if prov_pendientes > 0:
        print(f"\n  Tienes {prov_pendientes} proveedor(es) aún sin clasificar "
              f"(${gasto_pendiente:,.0f} MXN).")
        print(f"  Para subir la cobertura: clasifícalos en {NOMBRE_XLSX} y vuelve")
        print(f"  a correr este script subiendo {ARCHIVO_SAE} + el Excel actualizado.")
    else:
        print(f"\n  ✓ El 100% del gasto está categorizado.")
else:
    print(f"\n  Para mejorar la cobertura (actualmente {pct_cobertura:.1f}%):")
    print(f"  1. Abre {NOMBRE_XLSX}, clasifica los {prov_pendientes} proveedores pendientes.")
    print(f"  2. Vuelve a correr el script subiendo {ARCHIVO_SAE} + el Excel ya editado.")
print("═"*74)